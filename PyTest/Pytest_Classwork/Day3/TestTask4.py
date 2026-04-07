import pytest
import openpyxl
wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] = 'user'
ws['B1'] = 'password'

# wb.save('sample.xlsx')   # It will save inn the local directory

ws.append(['user1', '123'])
ws.append(['user2', '456'])
ws.append(['user3', '789'])
ws.append(['user4', '789'])
ws.append(['user5', '789'])

wb.save('sample.xlsx')  # To save the file

for row in ws.iter_rows(values_only=True):    # To iterate in the Excel Sheet
    print(row)